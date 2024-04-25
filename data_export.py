import sqlite3
from collections import namedtuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
from data_import import get_site_id_from_domain_name
from filters import check_for_node, is_high_priority
from openpyxl.worksheet.datavalidation import DataValidation

def get_all_sites():


    with open("sql/get_all_sites.sql", 'r') as file:
        sql_query = file.read()
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()
        cursor.execute(sql_query)
        results = cursor.fetchall()
        results = [result[0] for result in results]

        if not results:
            return []
        return results



def get_pdfs_by_site_name(site_name):

    with open("sql/get_pdf_reports_by_site_name.sql", 'r') as file:
        sql_query = file.read()
        formatted_query = sql_query.format(site_name=site_name)

        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()

        # Execute the SQL query
        cursor.execute(formatted_query)
        results = cursor.fetchall()

        # If there are no results, return an empty list
        if not results:
            return []

        # Get column names from the cursor
        col_names = [desc[0] for desc in cursor.description]
        # Create a namedtuple class with column names
        Row = namedtuple('Row', col_names)

        # Convert sqlite3.Row objects to namedtuples
        results = [Row(*row) for row in results]

        # Close the database connection
        conn.close()

        return results



def get_all_users_with_pdfs():
    with open("sql/get_all_users_with_pdf_files.sql", 'r') as file:
        sql_query = file.read()
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()
        cursor.execute(sql_query)
        results = cursor.fetchall()
        results = [result for result in results]

        if not results:
            return []
        return results





def get_site_failures(site_name):

    with open("sql/get_failures_by_site_id.sql", 'r') as file:
        sql_query = file.read()

        # need to get site id from site name

        site_id = get_site_id_from_domain_name(site_name.replace("-", "."))
        formatted_query = sql_query.format(site_id=site_id)
        conn = sqlite3.connect("drupal_pdfs.db")
        cursor = conn.cursor()

        cursor.execute(formatted_query)
        results = cursor.fetchall()

        # If there are no results, return an empty list
        if not results:
            return []


        # Get column names from the cursor
        col_names = [desc[0] for desc in cursor.description]
        # Create a namedtuple class with column names
        Row = namedtuple('Row', col_names)

        # Convert sqlite3.Row objects to namedtuples
        results = [Row(*row) for row in results]

        # Close the database connection
        conn.close()

        return results




from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font




def write_data_to_excel(data, failure_data, file_name="output.xlsx"):
    # Create a workbook
    wb = Workbook()

    # Create "Scanned PDFs" worksheet and set it as the active worksheet
    scanned_pdfs_ws = wb.create_sheet("Scanned PDFs", 0)
    wb.active = 0

    # Create "Failure" worksheet
    failure_ws = wb.create_sheet("Failure", 1)

    def add_data_to_scanned_pdfs(data, worksheet):
        if not data:
            print("No data to write to Excel.")
            return

        # Assuming all namedtuples have the same fields, get column names from the first item

        columns = list(data[0]._fields)
        print(columns)
        columns.remove('box_folder')
        columns.append("Errors/Page")
        columns.append("Low Priority")
        # Write the column headers to worksheet
        worksheet.append(columns)

        # Define the fill color for high priority rows
        red_fill = PatternFill(start_color='FF9999',  # This color code approximates "Red, Accent 2, Lighter 60%"
                               end_color='FF9999',
                               fill_type='solid')


        # Data validation for Low Priority column
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        worksheet.add_data_validation(dv)

        # Write the data rows to worksheet
        for item in data:

            item_list = list(item)
            high_priority = is_high_priority(item)

            if not check_for_node(item_list[1]):  # removes node urls
                item_list[0] = f'=HYPERLINK("{item[0]}", "{item[0]}")'
                item_list[1] = f'=HYPERLINK("{item[1]}", "{item[1]}")'
                item_list[3] = item[3][0:6]
                # Similar modifications as in your code...
                #item 6 = violations
                #item 7 = failed checks
                item_list[8] = "Yes" if item_list[8] == 1 else "No" #tagged
                #item 9 = pdf text type
                item_list[10] = "Yes" if item_list[10] == 1 else "No" # title set
                item_list[11] = "Yes" if item_list[11] == 1 else "No" # language set
                #item 12 = page_count
                # item_list[12] = "Yes" if item_list[12] == 1 else "No"
                item_list[13] = "Yes" if item_list[13] == 1 else "No"
                del item_list[14] # remove box.com link
                item_list.append(round(int(item[7]) / int(item[12])) if item[7] != 0 and item[12] !=0 else 0)
                item_list.append("No")

                # Append the modified list of values to the worksheet
                worksheet.append(item_list)
                dv.add(worksheet[f"{chr(65+len(columns)-1)}{worksheet._current_row}"])




                # If high_priority is True, apply red_fill to the entire row
                if high_priority:
                    for cell in worksheet[worksheet._current_row]:
                        cell.fill = red_fill

        # Apply font styles for hyperlink columns in worksheet
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=2, max_row=len(data) + 1):
            for cell in row:
                cell.font = Font(color='0563C1', underline='single')

        # Determine the Excel table range for worksheet
        table_range = f"A1:{chr(64 + len(columns))}{len(data) + 1}"

        # Create a table in worksheet
        table = Table(displayName="DataTable", ref=table_range)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=True)

        table.tableStyleInfo = style

        # Add the table to the worksheet
        worksheet.add_table(table)

    def add_data_to_failure(failure_data, worksheet):
        if not failure_data:
            print("No failure data to write to Excel.")
            return

        # Assuming all namedtuples have the same fields, get column names from the first item
        columns = list(failure_data[0]._fields)

        # Write the column headers to the worksheet
        worksheet.append(columns)

        # Write the data rows to the worksheet
        for item in failure_data:
            # Convert named tuple to a list
            item_list = list(item)

            # Append the list of values to the worksheet
            worksheet.append(item_list)

        # Determine the Excel table range for the worksheet
        table_range = f"A1:{chr(64 + len(columns))}{len(failure_data) + 1}"

        # Create a table in the worksheet
        table = Table(displayName="FailureDataTable", ref=table_range)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=True)

        table.tableStyleInfo = style

        # Add the table to the worksheet
        worksheet.add_table(table)

    # Call the sub-function to add data to the "Scanned PDFs" worksheet
    add_data_to_scanned_pdfs(data, scanned_pdfs_ws)
    add_data_to_failure(failure_data, failure_ws)

    # Placeholder call for adding data to the "Failure" worksheet
    # Uncomment and use when implementation is added
    # add_data_to_failure(failure, failure_ws)

    # Save the Excel file
    wb.save(file_name)
    print(f"Data written to {file_name} with a table format.")


