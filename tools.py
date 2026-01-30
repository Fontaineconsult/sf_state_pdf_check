
import os
import re
import html
import shutil
import zipfile
from datetime import datetime

import requests
from urllib.parse import unquote
from openpyxl import load_workbook

DEFAULT_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

from conformance_checker import loop_through_files_in_folder
from data_export import get_pdf_reports_by_site_name
from data_import import get_site_id_by_domain_name, mark_pdf_as_removed
from sf_state_pdf_scan.sf_state_pdf_scan.box_handler import download_from_box, box_share_pattern_match, normalize_box_url
from set_env import get_box_path, get_database_path


pdf_sites_folder = get_box_path('pdf_scans')
scans_output = get_box_path('pdf_scans') + "\\{}"


def delete_scans_files(root_folder):
    # Walk through the directory tree
    for current_path, dirs, files in os.walk(root_folder):
        for file in files:
            # Check if the file name contains the substring
            if "_scans.xlsx" in file:
                file_path = os.path.join(current_path, file)
                try:
                    os.remove(file_path)
                    print(f"Deleted: {file_path}")
                except Exception as e:
                    print(f"Error deleting {file_path}: {e}")


def remove_timestamps_from_parent_urls():
    conn = sqlite3.connect(get_database_path())
    cursor = conn.cursor()
    pdfs = cursor.execute("SELECT * FROM drupal_pdf_files").fetchall()

    for each in pdfs:
        original_url = each[2]
        print(each[0])
        # Remove everything after the first space (if any)
        cleaned_url = original_url.split(" ")[0]

        # If the URL has been changed, update the record
        if cleaned_url != original_url:
            print(f"Updating: {original_url} -> {cleaned_url}")
            # Assuming 'id' is the primary key and is stored in the first column.
            cursor.execute("UPDATE drupal_pdf_files SET parent_uri = ? WHERE id = ?", (cleaned_url, each[0]))

    conn.commit()
    conn.close()

def strip_trailing_items_from_pdf_urls():
    conn = sqlite3.connect(get_database_path())
    cursor = conn.cursor()
    pdfs = cursor.execute("SELECT * FROM drupal_pdf_files").fetchall()

    for each in pdfs:
        original_url = each[1]
        # print(each[1])
        # Remove everything after the first space (if any)
        cleaned_url = original_url.split(" ")[0]

        # If the URL has been changed, update the record
        if cleaned_url != original_url:
            print(f"Updating: {original_url} -> {cleaned_url}")
            # Assuming 'id' is the primary key and is stored in the first column.
            cursor.execute("UPDATE drupal_pdf_files SET pdf_uri = ? WHERE id = ?", (cleaned_url, each[0]))

    conn.commit()
    conn.close()

import sqlite3

def delete_duplicate_entries():
    """
    For each site in the drupal_site table, this function reads the SQL query from
    delete_duplicates.sql (which deletes duplicate PDF entries for a given site by keeping only the oldest scan),
    substitutes the {site_name} placeholder with the actual domain name, and executes the query.

    The site names are retrieved by executing the SQL query stored in get_all_sites.sql.
    """
    # Connect to the SQLite database.
    conn = sqlite3.connect(get_database_path())
    cursor = conn.cursor()

    # Read the get_all_sites.sql file, which contains:
    # select domain_name from drupal_site;
    with open("sql/get_all_sites.sql", "r") as f:
        get_sites_query = f.read().strip()

    # Execute the query to fetch all site domain names.
    cursor.execute(get_sites_query)
    sites = cursor.fetchall()  # Each row is a tuple (domain_name,)

    # Read the delete_duplicates.sql file which contains our delete query template.
    with open("sql/delete_duplicates.sql", "r") as f:
        delete_query_template = f.read()

    # Loop through each site and execute the delete query with the proper substitution.
    for site in sites:
        domain_name = site[0]
        formatted_query = delete_query_template.replace("{site_name}", domain_name)
        print(f"Deleting duplicate entries for site: {domain_name}")
        cursor.executescript(formatted_query)
        conn.commit()
        print(f"Finished processing site: {domain_name}")

    # Close the database connection.
    conn.close()


def download_all_dprc_will_remediate_pdfs_by_site(site_name):
    box_folder = os.path.join(get_box_path('pdf_scans'), site_name)
    box_temp_folder = os.path.join(box_folder, 'temp')
    xlsx_file = os.path.join(box_folder, f"{site_name.split('-')[0]}-pdf-scans.xlsx")

    if not os.path.exists(box_folder):
        print(f"File does not exist: {box_folder}")
        return
    print(f"Found folder: {box_folder}")

    # Compile hyperlink regex
    hyperlink_pattern = re.compile(r'HYPERLINK\("([^"]+)"')

    # Load workbook and sheet
    workbook = load_workbook(xlsx_file)
    sheet = workbook['Scanned PDFs']
    print(f"Processing sheet: {sheet.title}")

    # Prepare temp folder
    os.makedirs(box_temp_folder, exist_ok=True)

    # Download loop
    for row in sheet.iter_rows(min_row=1, max_col=17, max_row=sheet.max_row):
        link = row[0].value
        high_priority = row[15].value
        dprc_remediation = row[16].value

        if dprc_remediation == "Yes" and isinstance(link, str) and link.startswith('=HYPERLINK'):
            match = hyperlink_pattern.search(link)
            if not match:
                continue

            first_url = match.group(1)
            print(first_url, high_priority, dprc_remediation)

            if box_share_pattern_match(first_url):
                print("Downloading file from box…")
                download_from_box(first_url, box_temp_folder)
            else:
                print("Downloading file from URL…")
                response = requests.get(first_url, headers=DEFAULT_HEADERS, stream=True)
                if response.status_code == 200:
                    raw_name = os.path.basename(first_url)
                    unescaped = html.unescape(raw_name)
                    file_name = unquote(unescaped)
                    file_path = os.path.join(box_temp_folder, file_name)
                    with open(file_path, "wb") as file:
                        for chunk in response.iter_content(chunk_size=8192):
                            file.write(chunk)

    # --- ZIP and cleanup ---
    # Define the ZIP filename
    zip_name = f"{site_name}-pdf-scans.zip"
    zip_path = os.path.join(box_folder, zip_name)

    print(f"Creating ZIP archive: {zip_path}")
    with zipfile.ZipFile(zip_path, 'w', ZIP_DEFLATED:=zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(box_temp_folder):
            for fname in files:
                if fname.lower().endswith('.pdf'):
                    full_path = os.path.join(root, fname)
                    # Store PDFs at the root of the archive, no temp/ prefix
                    arcname = os.path.relpath(full_path, box_temp_folder)
                    zipf.write(full_path, arcname)

    print("ZIP archive created successfully.")

    # Remove the temp folder
    print(f"Removing temporary folder: {box_temp_folder}")
    shutil.rmtree(box_temp_folder)
    print("Temporary folder deleted.")


def get_all_folders_by_date_modified(folder_path, date_modified):
    """
    Returns a list of all folder names (not full paths) in the specified directory that were modified after the given date.

    Parameters:
        folder_path (str): The path to the directory to search.
        date_modified (str): The date (YYYY-MM-DD or MM/DD/YYYY) to compare against.

    Returns:
        list: A list of folder names that were modified after the specified date.
    """
    # Try parsing date in multiple formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            date_obj = datetime.strptime(date_modified, fmt)
            break
        except ValueError:
            continue
    else:
        raise ValueError("date_modified must be in YYYY-MM-DD or MM/DD/YYYY format")

    modified_folders = []
    for root, dirs, _ in os.walk(folder_path):
        for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            if os.path.getmtime(dir_path) > date_obj.timestamp():
                modified_folders.append(dir_name)
    return modified_folders





# print('\n'.join(get_all_folders_by_date_modified(
#     r"C:\Users\913678186\Box\ATI\PDF Accessibility\SF State Website PDF Scans",
#     "06/17/2025")))

# download_all_dprc_will_remediate_pdfs_by_site('cob-sfsu-edu')

def add_passed_contains_accessible_in_title_column():
    """
    Add the passed_contains_accessible_in_title column to drupal_pdf_files table.
    """
    conn = sqlite3.connect(get_database_path())
    cursor = conn.cursor()

    cursor.execute("""
        ALTER TABLE drupal_pdf_files
        ADD COLUMN passed_contains_accessible_in_title boolean NOT NULL DEFAULT false
    """)

    print("Added column 'passed_contains_accessible_in_title' to drupal_pdf_files table.")

    conn.commit()
    conn.close()


def mark_pdfs_with_accessible_in_title_as_passed():
    """
    Mark PDFs as passed if their pdf_uri contains 'accessible' (case-insensitive).
    Updates the passed_contains_accessible_in_title column to 1 (true) for matching PDFs.
    """
    conn = sqlite3.connect(get_database_path())
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE drupal_pdf_files
        SET passed_contains_accessible_in_title = 1
        WHERE LOWER(pdf_uri) LIKE '%accessible%'
    """)

    updated_count = cursor.rowcount
    print(f"Marked {updated_count} PDFs with 'accessible' in title as passed.")

    conn.commit()
    conn.close()



def mark_pdfs_as_removed(site_folders):
    """
    Compare the raw pdf scrape URLS and Parent with current PDFS and mark current PDFS as removed if they are not in the raw scrape.
    Box URLs are normalized before comparison to handle sfsu.box.com vs sfsu.app.box.com variations.
    :return:
    """
    raw_pdf_scan_set = set()

    for folder in os.listdir(site_folders):

        domain_id = get_site_id_by_domain_name(folder)

        site_pdfs = get_pdf_reports_by_site_name(folder.replace("-", "."))

        # Build existing set with normalized Box URLs, but keep original for marking
        existing_pdfs_map = {(normalize_box_url(pdf.pdf_uri), pdf.parent_uri): (pdf.pdf_uri, pdf.parent_uri) for pdf in site_pdfs}
        existing_pdfs_set = set(existing_pdfs_map.keys())

        if domain_id is not None:
            pdf_locations = loop_through_files_in_folder(os.path.join(site_folders, folder))

            # If no scan file exists or is empty, skip this site to avoid marking all PDFs as removed
            if not pdf_locations:
                print(f"No scanned_pdfs.txt found or empty for {folder}, skipping removal check")
                raw_pdf_scan_set.clear()
                existing_pdfs_set.clear()
                existing_pdfs_map.clear()
                continue

            for file in pdf_locations:
                file_split = file.split(' ', 1)  # Splits at the first space
                file_url = file_split[0]
                loc = file_split[1].split(" ")[0]
                # Normalize Box URLs for comparison
                raw_pdf_scan_set.add((normalize_box_url(file_url), loc))

            missing_pdfs = existing_pdfs_set.difference(raw_pdf_scan_set)
            if missing_pdfs:
                for normalized_key in missing_pdfs:
                    # Use original URL from map for marking
                    original_pdf_uri, original_parent_uri = existing_pdfs_map[normalized_key]
                    mark_pdf_as_removed(original_pdf_uri, original_parent_uri)

        raw_pdf_scan_set.clear()
        existing_pdfs_set.clear()
        existing_pdfs_map.clear()


def mark_single_site_pdfs_as_removed(site_folder):
    """
    Compare the raw pdf scrape URLs and Parent with current PDFs for a single site
    and mark current PDFs as removed if they are not in the raw scrape.
    Box URLs are normalized before comparison to handle sfsu.box.com vs sfsu.app.box.com variations.

    Parameters:
        site_folder (str): Path to the single site folder to process
    """
    folder_name = os.path.basename(site_folder)
    domain_name = folder_name.replace("-", ".")

    domain_id = get_site_id_by_domain_name(folder_name)
    if domain_id is None:
        print(f"Could not find domain ID for {folder_name}")
        return

    site_pdfs = get_pdf_reports_by_site_name(domain_name)
    # Build existing set with normalized Box URLs, but keep original for marking
    existing_pdfs_map = {(normalize_box_url(pdf.pdf_uri), pdf.parent_uri): (pdf.pdf_uri, pdf.parent_uri) for pdf in site_pdfs}
    existing_pdfs_set = set(existing_pdfs_map.keys())

    raw_pdf_scan_set = set()
    pdf_locations = loop_through_files_in_folder(site_folder)

    # If no scan file exists or is empty, skip removal to avoid marking all PDFs as removed
    if not pdf_locations:
        print(f"No scanned_pdfs.txt found or empty for {domain_name}, skipping removal check")
        return

    for file in pdf_locations:
        file_split = file.split(' ', 1)
        file_url = file_split[0]
        loc = file_split[1].split(" ")[0]
        # Normalize Box URLs for comparison
        raw_pdf_scan_set.add((normalize_box_url(file_url), loc))

    missing_pdfs = existing_pdfs_set.difference(raw_pdf_scan_set)
    if missing_pdfs:
        for normalized_key in missing_pdfs:
            # Use original URL from map for marking
            original_pdf_uri, original_parent_uri = existing_pdfs_map[normalized_key]
            mark_pdf_as_removed(original_pdf_uri, original_parent_uri)

    print(f"Marked {len(missing_pdfs)} PDFs as removed for {domain_name}")


if __name__ == "__main__":
    mark_pdfs_as_removed(pdf_sites_folder)

